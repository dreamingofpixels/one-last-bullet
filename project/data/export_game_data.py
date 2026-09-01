"""
Export game_data.xlsx to game_data.json using the "schema" sheet for column types.
Sheets listed in the schema are exported as lists under the same keys in the output JSON.
"""
from pathlib import Path
import json
from openpyxl import load_workbook


def main() -> None:
    data_dir = Path(__file__).resolve().parent
    xlsx_path = data_dir / "game_data.xlsx"
    out_path = data_dir / "game_data.json"

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # 1. Read schema: sheet, column, type
    schema = wb["schema"]
    type_map: dict[tuple[str, str], str] = {}
    data_sheet_names: set[str] = set()
    for row in schema.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sheet_name = str(row[0]).strip()
        col_name = str(row[1]).strip() if row[1] is not None else ""
        type_name = str(row[2]).strip().lower() if row[2] is not None else "string"
        if not col_name:
            continue
        type_map[(sheet_name, col_name)] = type_name
        if sheet_name != "schema":
            data_sheet_names.add(sheet_name)

    # 2. Export each data sheet
    out: dict[str, list] = {}
    for sheet_name in sorted(data_sheet_names):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[sheet_name] = []
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        records = []
        for row in rows[1:]:
            record = {}
            for i, raw in enumerate(row):
                if i >= len(headers) or not headers[i]:
                    continue
                key = headers[i]
                type_name = type_map.get((sheet_name, key), "string")
                value = _coerce(raw, type_name)
                # Robustness: allow string IDs even if the schema says "int".
                if key == "id" and value is None and raw is not None:
                    raw_s = str(raw).strip()
                    if raw_s != "":
                        value = raw_s
                record[key] = value

            if not record or all(v is None for v in record.values()):
                continue
            id_v = record.get("id")
            if id_v is None or (isinstance(id_v, str) and id_v.strip() == ""):
                continue
            records.append(record)
        out[sheet_name] = records

    wb.close()

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f"Wrote {out_path}")


def _coerce(value, type_name: str):
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if type_name == "int":
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(float(value))
        except (ValueError, TypeError):
            return None
    if type_name == "float":
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    if type_name == "boolean":
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        s = str(value).strip().lower()
        if s in ("true", "1", "yes"):
            return True
        if s in ("false", "0", "no"):
            return False
        return None
    return str(value).strip() if value is not None else ""


if __name__ == "__main__":
    main()
